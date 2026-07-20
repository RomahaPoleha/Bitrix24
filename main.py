import requests
import json
import time
import os
import uuid
import socket
import base64
import hashlib
from datetime import datetime, timedelta, date as date_type, time as time_type
from zoneinfo import ZoneInfo
import tkinter as tk
from tkinter import messagebox, simpledialog
from cryptography.fernet import Fernet
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# === 1. ПУТИ И КОНСТАНТЫ ===
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
WEBHOOK_FILE = os.path.join(SCRIPT_DIR, "webhook.enc")
HOLIDAYS_FILE = os.path.join(SCRIPT_DIR, "holidays.json")

# Секретная соль (можно изменить на свою случайную строку)
# Это усилит защиту — даже при одинаковых данных машины ключ будет уникальным для твоей версии программы
SECRET_SALT = b"ModernSchool_2026_B24_Export_Tool_v1"


# === 2. ШИФРОВАНИЕ ===

def get_machine_id():
    """Собирает уникальные данные компьютера для генерации ключа"""
    machine_data = []

    # Имя пользователя
    try:
        machine_data.append(os.getlogin())
    except:
        machine_data.append("unknown_user")

    # Имя компьютера
    try:
        machine_data.append(socket.gethostname())
    except:
        machine_data.append("unknown_host")

    # MAC-адрес
    try:
        mac = ':'.join(['{:02x}'.format((uuid.getnode() >> i) & 0xff)
                        for i in range(0, 2 * 6, 8)][::-1])
        machine_data.append(mac)
    except:
        machine_data.append("unknown_mac")

    return "|".join(machine_data).encode("utf-8")


def get_encryption_key():
    """Генерирует ключ шифрования на основе данных машины и соли"""
    machine_id = get_machine_id()
    # Смешиваем machine_id с солью через SHA256
    combined = machine_id + SECRET_SALT
    key_hash = hashlib.sha256(combined).digest()
    # Fernet требует ключ в base64 (32 байта → 44 символа)
    return base64.urlsafe_b64encode(key_hash)


def encrypt_url(url):
    """Шифрует URL вебхука"""
    key = get_encryption_key()
    f = Fernet(key)
    return f.encrypt(url.encode("utf-8")).decode("utf-8")


def decrypt_url(encrypted):
    """Расшифровывает URL вебхука"""
    key = get_encryption_key()
    f = Fernet(key)
    return f.decrypt(encrypted.encode("utf-8")).decode("utf-8")


# === 3. ЗАГРУЗКА/СОХРАНЕНИЕ КЛЮЧА ===

def validate_webhook(url):
    """Проверяет, что URL похож на вебхук Битрикс24"""
    url = url.strip()
    if not url.startswith("https://"):
        return False, "URL должен начинаться с https://"
    if "/rest/" not in url:
        return False, "URL должен содержать /rest/"
    return True, "OK"


def test_webhook(url):
    """Делает тестовый запрос к API для проверки валидности ключа"""
    try:
        if not url.endswith("/"):
            url += "/"
        response = requests.get(f"{url}app.info.json", timeout=10)
        if response.status_code == 200:
            data = response.json()
            if "result" in data:
                return True, "Ключ работает"
        return False, f"API вернул статус {response.status_code}"
    except Exception as e:
        return False, f"Ошибка подключения: {e}"


def show_webhook_dialog():
    """Показывает окно ввода вебхука"""
    dialog = tk.Toplevel()
    dialog.title("Настройка подключения")
    dialog.geometry("550x250")
    dialog.resizable(False, False)
    dialog.transient()  # Привязываем к главному окну
    dialog.grab_set()  # Модальное окно

    # Делаем диалог поверх всех окон
    dialog.attributes("-topmost", True)

    tk.Label(dialog, text="Введите URL вебхука Битрикс24",
             font=("Arial", 12, "bold")).pack(pady=15)

    tk.Label(dialog, text="Где взять: Приложения → Разработчикам → Другое → Входящий вебхук",
             fg="gray").pack()

    frame_entry = tk.Frame(dialog)
    frame_entry.pack(pady=10)

    entry = tk.Entry(frame_entry, width=60, font=("Arial", 10), show="*")
    entry.pack(side=tk.LEFT, padx=5)

    # Кнопка показать/скрыть
    def toggle_visibility():
        if entry.cget("show") == "*":
            entry.config(show="")
            btn_show.config(text="Скрыть")
        else:
            entry.config(show="*")
            btn_show.config(text="Показать")

    btn_show = tk.Button(frame_entry, text="Показать", command=toggle_visibility, width=8)
    btn_show.pack(side=tk.LEFT, padx=5)

    result = {"url": None}

    def on_save():
        url = entry.get().strip()

        valid, msg = validate_webhook(url)
        if not valid:
            messagebox.showerror("Ошибка формата", msg, parent=dialog)
            return

        # Проверяем ключ
        status_label.config(text="Проверяю ключ...", fg="blue")
        dialog.update()

        success, msg = test_webhook(url)
        if not success:
            status_label.config(text=f"Ошибка: {msg}", fg="red")
            messagebox.showerror("Ошибка подключения",
                                 f"Не удалось подключиться к Битрикс24:\n{msg}\n\n"
                                 f"Проверьте URL и попробуйте снова.",
                                 parent=dialog)
            return

        result["url"] = url
        dialog.destroy()

    def on_cancel():
        dialog.destroy()

    frame_btns = tk.Frame(dialog)
    frame_btns.pack(pady=10)

    tk.Button(frame_btns, text="Сохранить", command=on_save,
              bg="#4472C4", fg="white", font=("Arial", 10, "bold"),
              width=12, height=1).pack(side=tk.LEFT, padx=5)
    tk.Button(frame_btns, text="Отмена", command=on_cancel,
              width=12, height=1).pack(side=tk.LEFT, padx=5)

    status_label = tk.Label(dialog, text="", fg="gray")
    status_label.pack(pady=5)

    # Enter для сохранения
    dialog.bind("<Return>", lambda e: on_save())

    # Центрируем относительно главного окна
    dialog.update_idletasks()
    dialog.wait_window()

    return result["url"]


def save_webhook(url):
    """Шифрует и сохраняет URL в файл"""
    try:
        encrypted = encrypt_url(url)
        with open(WEBHOOK_FILE, "w", encoding="utf-8") as f:
            f.write(encrypted)
        print("Ключ успешно сохранён (в зашифрованном виде)")
        return True
    except Exception as e:
        print(f"Ошибка сохранения ключа: {e}")
        return False


def load_webhook():
    """Загружает вебхук. Если файла нет — просит ввести."""
    # Если файла нет — показываем окно ввода
    if not os.path.exists(WEBHOOK_FILE):
        print("Файл с ключом не найден. Запрашиваю ввод...")
        # Нужно создать скрытое главное окно, чтобы диалог работал
        temp_root = tk.Tk()
        temp_root.withdraw()  # Скрываем главное окно
        temp_root.attributes("-topmost", True)

        url = show_webhook_dialog()
        temp_root.destroy()

        if not url:
            print("Ввод ключа отменён. Завершение работы.")
            exit()

        if save_webhook(url):
            return url
        else:
            print("Не удалось сохранить ключ. Завершение работы.")
            exit()

    # Файл есть — читаем и расшифровываем
    try:
        with open(WEBHOOK_FILE, "r", encoding="utf-8") as f:
            encrypted = f.read().strip()

        url = decrypt_url(encrypted)

        # Нормализуем URL
        if not url.endswith("/"):
            url += "/"

        print("Ключ успешно загружен и расшифрован")
        return url

    except Exception as e:
        print(f"Ошибка расшифровки ключа: {e}")
        print("Возможно, файл повреждён или вы перенесли программу на другой компьютер.")

        # Предлагаем ввести заново
        temp_root = tk.Tk()
        temp_root.withdraw()
        temp_root.attributes("-topmost", True)

        answer = messagebox.askyesno(
            "Ошибка ключа",
            "Не удалось расшифровать сохранённый ключ.\n"
            "Возможно, программа перенесена на другой компьютер.\n\n"
            "Ввести ключ заново?"
        )

        if answer:
            url = show_webhook_dialog()
            temp_root.destroy()

            if url and save_webhook(url):
                return url

        temp_root.destroy()
        print("Завершение работы.")
        exit()


# === 4. ОСНОВНАЯ КОНФИГУРАЦИЯ ===

WEBHOOK_URL = load_webhook()
PORTAL_URL = WEBHOOK_URL.split("/rest/")[0]

USER_TZ = ZoneInfo("Asia/Vladivostok")

WORK_START = time_type(9, 0)
WORK_END = time_type(18, 0)

WORK_INTERVALS = [
    (time_type(9, 0), time_type(11, 0)),
    (time_type(11, 15), time_type(13, 0)),
    (time_type(14, 0), time_type(16, 0)),
    (time_type(16, 15), time_type(18, 0)),
]

LONG_TASK_THRESHOLD = 4 * 60


# === 5. ПРАЗДНИЧНЫЕ ДНИ (без изменений) ===

def get_hardcoded_holidays(year):
    return [
        f"{year}-01-01", f"{year}-01-02", f"{year}-01-03", f"{year}-01-04",
        f"{year}-01-05", f"{year}-01-06", f"{year}-01-07", f"{year}-01-08",
        f"{year}-02-23", f"{year}-03-08", f"{year}-05-01", f"{year}-05-09",
        f"{year}-06-12", f"{year}-11-04"
    ]


def load_holidays():
    if not os.path.exists(HOLIDAYS_FILE):
        return {"last_updated": None, "years": {}}
    try:
        with open(HOLIDAYS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"last_updated": None, "years": {}}


def save_holidays(data):
    try:
        with open(HOLIDAYS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Ошибка сохранения {HOLIDAYS_FILE}: {e}")


def fetch_holidays_from_api(year):
    try:
        url = f"https://isdayoff.ru/api/getdata?year={year}&cc=ru&ru"
        response = requests.get(url, timeout=10)

        if response.status_code != 200:
            return get_hardcoded_holidays(year)

        data_str = response.text.strip()
        if len(data_str) < 365:
            return get_hardcoded_holidays(year)

        holidays = []
        start_date = datetime(year, 1, 1).date()
        for i, day_type in enumerate(data_str):
            if day_type == '1':
                holiday_date = start_date + timedelta(days=i)
                holidays.append(holiday_date.strftime("%Y-%m-%d"))

        return holidays
    except Exception:
        return get_hardcoded_holidays(year)


def update_holidays_if_needed():
    holidays_data = load_holidays()
    current_year = datetime.now().year
    last_updated = holidays_data.get("last_updated")
    years_data = holidays_data.get("years", {})

    if str(current_year) in years_data and last_updated:
        try:
            last_updated_date = datetime.strptime(last_updated, "%Y-%m-%d").date()
            if (datetime.now().date() - last_updated_date).days < 7:
                return years_data
        except ValueError:
            pass

    holidays_list = fetch_holidays_from_api(current_year)
    if holidays_list:
        years_data[str(current_year)] = holidays_list
    else:
        years_data[str(current_year)] = get_hardcoded_holidays(current_year)

    holidays_data["years"] = years_data
    holidays_data["last_updated"] = datetime.now().strftime("%Y-%m-%d")
    save_holidays(holidays_data)
    return years_data


HOLIDAYS_CACHE = None


def get_holidays():
    global HOLIDAYS_CACHE
    if HOLIDAYS_CACHE is None:
        HOLIDAYS_CACHE = update_holidays_if_needed()
    return HOLIDAYS_CACHE


def is_holiday(d):
    holidays = get_holidays()
    year_str = str(d.year)
    if year_str not in holidays:
        return False
    date_str = d.strftime("%Y-%m-%d")
    return date_str in holidays[year_str]


# === 6. РАСЧЁТ РАБОЧЕГО ВРЕМЕНИ ===

def get_task_url(task_id, user_id):
    return f"{PORTAL_URL}/company/personal/user/{user_id}/tasks/task/view/{task_id}/"


def parse_date(date_str):
    if not date_str: return None
    try:
        return datetime.fromisoformat(date_str)
    except:
        return None


def format_minutes(minutes):
    if minutes <= 0: return "0 мин"
    h, m = minutes // 60, minutes % 60
    return f"{h} ч {m} мин" if h > 0 else f"{m} мин"


def is_working_day(d):
    if d.weekday() >= 5: return False
    if is_holiday(d): return False
    return True


def time_to_minutes(t): return t.hour * 60 + t.minute


def working_minutes_in_interval(start_t, end_t, interval_start, interval_end):
    s = max(time_to_minutes(start_t), time_to_minutes(interval_start))
    e = min(time_to_minutes(end_t), time_to_minutes(interval_end))
    return max(0, e - s)


def working_minutes_in_day(d, start_t=None, end_t=None):
    if not is_working_day(d): return 0
    if start_t is None: start_t = WORK_START
    if end_t is None: end_t = WORK_END
    if time_to_minutes(start_t) >= time_to_minutes(WORK_END): return 0
    if time_to_minutes(end_t) <= time_to_minutes(WORK_START): return 0
    start_t, end_t = max(start_t, WORK_START), min(end_t, WORK_END)
    return sum(working_minutes_in_interval(start_t, end_t, i_start, i_end)
               for i_start, i_end in WORK_INTERVALS)


def calculate_working_minutes(start_dt, end_dt):
    start_vl, end_vl = start_dt.astimezone(USER_TZ), end_dt.astimezone(USER_TZ)
    if start_vl >= end_vl: return 0
    start_date, end_date = start_vl.date(), end_vl.date()
    if start_date == end_date:
        return working_minutes_in_day(start_date, start_vl.time(), end_vl.time())
    total = working_minutes_in_day(start_date, start_vl.time(), WORK_END)
    current = start_date + timedelta(days=1)
    while current < end_date:
        total += working_minutes_in_day(current)
        current += timedelta(days=1)
    total += working_minutes_in_day(end_date, WORK_START, end_vl.time())
    return total


# === 7. ВЫГРУЗКА ИЗ БИТРИКС24 ===

def get_all_tasks(date_from, date_to, hashtag=None):
    all_tasks, start = [], 0
    date_from_str = datetime.combine(date_from, time_type.min).replace(tzinfo=USER_TZ).astimezone(
        ZoneInfo("UTC")).strftime("%Y-%m-%dT%H:%M:%S+00:00")
    date_to_str = datetime.combine(date_to, time_type.max).replace(tzinfo=USER_TZ).astimezone(ZoneInfo("UTC")).strftime(
        "%Y-%m-%dT%H:%M:%S+00:00")

    task_filter = {">CREATED_DATE": date_from_str, "<CREATED_DATE": date_to_str}
    if hashtag: task_filter["TAG"] = hashtag.lstrip("#")

    while True:
        try:
            response = requests.post(f"{WEBHOOK_URL}tasks.task.list.json",
                                     json={"filter": task_filter,
                                           "select": ["ID", "TITLE", "STATUS", "RESPONSIBLE_ID", "CLOSED_DATE",
                                                      "CREATED_DATE", "TAGS"],
                                           "start": start}, timeout=30)
        except Exception as e:
            print(f"Ошибка подключения: {e}");
            break
        if response.status_code != 200:
            print(f"HTTP ошибка: {response.status_code}");
            break
        try:
            result = response.json()
        except:
            break
        if "error" in result:
            print(f"Ошибка API: {result['error']}");
            break

        tasks = result.get("result", {}).get("tasks", [])
        if not tasks: break
        all_tasks.extend(tasks)
        print(f"Загружено {len(all_tasks)} задач...")

        next_start = result.get("next")
        if not next_start: break
        start = next_start
        time.sleep(0.2)
    return all_tasks


def is_task_closed(task): return str(task.get("status")) == "5"


# === 8. СОХРАНЕНИЕ В EXCEL ===

def save_to_excel(tasks, filename="tasks_report.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Задачи"
    headers = ["Название", "Дата создания", "Дата закрытия", "Время выполнения", "Ссылка"]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for task in tasks:
        created = parse_date(task.get("createdDate"))
        closed = parse_date(task.get("closedDate"))
        created_vl = created.astimezone(USER_TZ) if created else None
        closed_vl = closed.astimezone(USER_TZ) if closed else None

        if is_task_closed(task) and closed_vl:
            working_mins = calculate_working_minutes(created, closed)
            is_long = working_mins > LONG_TASK_THRESHOLD
            row = [
                task.get("title", ""),
                created_vl.strftime("%d.%m.%Y %H:%M"),
                closed_vl.strftime("%d.%m.%Y %H:%M"),
                working_mins,
                get_task_url(task["id"], task.get("responsibleId", "1"))
            ]
        else:
            is_long = False
            row = [
                task.get("title", ""),
                created_vl.strftime("%d.%m.%Y %H:%M") if created_vl else "N/A",
                "В процессе",
                "Выполняется",
                get_task_url(task["id"], task.get("responsibleId", "1"))
            ]

        ws.append(row)
        if is_long:
            for cell in ws[ws.max_row]:
                cell.fill = yellow_fill

    for col, width in zip(["A", "B", "C", "D", "E"], [50, 20, 20, 18, 70]):
        ws.column_dimensions[col].width = width

    # Форматируем колонку "Время выполнения" как число
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            if isinstance(cell.value, int):
                cell.number_format = '0'

    save_path = os.path.join(SCRIPT_DIR, filename)
    wb.save(save_path)
    print(f"Сохранено в {save_path}")


# === 9. ФУНКЦИЯ ЭКСПОРТА ===

def parse_date_input(text):
    try:
        return datetime.strptime(text.strip(), "%d.%m.%Y").date()
    except:
        return None


def start_export():
    date_from = parse_date_input(entry_from.get())
    date_to = parse_date_input(entry_to.get())
    hashtag = entry_hashtag.get().strip()

    if not date_from or not date_to:
        messagebox.showerror("Ошибка", "Введите даты в формате ДД.ММ.ГГГГ");
        return
    if date_from > date_to:
        messagebox.showerror("Ошибка", "Дата 'С' не может быть позже даты 'По'");
        return

    print(f"\nПериод: {date_from} - {date_to}")
    if hashtag: print(f"Хештег: {hashtag}")

    all_tasks = get_all_tasks(date_from, date_to, hashtag if hashtag else None)
    print(f"Всего задач: {len(all_tasks)}")

    if all_tasks:
        save_to_excel(all_tasks)
        messagebox.showinfo("Готово", f"Выгружено {len(all_tasks)} задач.\nФайл: tasks_report.xlsx")
    else:
        messagebox.showinfo("Готово", "Нет задач, подходящих под условия")


def change_webhook():
    """Функция для смены ключа через меню"""
    url = show_webhook_dialog()
    if url:
        if save_webhook(url):
            messagebox.showinfo("Готово", "Ключ успешно обновлён.\nПерезапустите программу для применения.")
        else:
            messagebox.showerror("Ошибка", "Не удалось сохранить ключ.")


# === 10. GUI ===

root = tk.Tk()
root.title("Тест")
root.geometry("500x420")

# Меню с возможностью смены ключа
menu_bar = tk.Menu(root)
settings_menu = tk.Menu(menu_bar, tearoff=0)
settings_menu.add_command(label="Сменить ключ доступа", command=change_webhook)
settings_menu.add_separator()
settings_menu.add_command(label="Выход", command=root.quit)
menu_bar.add_cascade(label="Настройки", menu=settings_menu)
root.config(menu=menu_bar)

tk.Label(root, text="Выберите период", font=("Arial", 14, "bold")).pack(pady=10)

frame_from = tk.Frame(root)
frame_from.pack(pady=5)
tk.Label(frame_from, text="С даты (ДД.ММ.ГГГГ): ", width=22, anchor="w").pack(side=tk.LEFT)
entry_from = tk.Entry(frame_from, width=15, font=("Arial", 11))
entry_from.pack(side=tk.LEFT)
entry_from.insert(0, "01.01.2026")

frame_to = tk.Frame(root)
frame_to.pack(pady=5)
tk.Label(frame_to, text="По дату (ДД.ММ.ГГГГ): ", width=22, anchor="w").pack(side=tk.LEFT)
entry_to = tk.Entry(frame_to, width=15, font=("Arial", 11))
entry_to.pack(side=tk.LEFT)
entry_to.insert(0, datetime.now().strftime("%d.%m.%Y"))

frame_hashtag = tk.Frame(root)
frame_hashtag.pack(pady=5)
tk.Label(frame_hashtag, text="Хештег (необязательно): ", width=22, anchor="w").pack(side=tk.LEFT)
entry_hashtag = tk.Entry(frame_hashtag, width=15, font=("Arial", 11))
entry_hashtag.pack(side=tk.LEFT)

tk.Button(root, text="Выгрузить задачи", command=start_export,
          bg="#4472C4", fg="white", font=("Arial", 12, "bold"),
          padx=20, pady=10).pack(pady=20)

tk.Label(root, text="ООО Современная Школа", fg="gray", justify="center").pack()

# Инициализация кэша праздников
_ = get_holidays()

root.mainloop()